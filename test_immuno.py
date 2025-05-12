import pandas as pd
import numpy as np
import csv
from Bio import SeqIO
import os
import subprocess
import sys
import re
import ast
import papermill
import logging
import argparse
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

#functions 

#fn for creating files from a list and file_name
def create_files(pep, file_name):

    ofile = open(file_name + ".fasta", "w")

    for i in range(len(pep)):
        ofile.write('>' + '\n' + pep[i] + '\n')
    ofile.close()

    # peptide list for parsing
    ofile = open(file_name + '_2.fasta', "w")

    for i in range(len(pep)):
        ofile.write(pep[i] + '\n')
    ofile.close()
    logger.info(f"\tfiles {file_name}, {file_name}_2 created")
    return True

#fn for parsing and categorizing blast output
def parse_categorize(database_fasta, blast_out, fasta2, SAAV=False): #requires inputs with file extension

    input1= SeqIO.parse(database_fasta,"fasta") # blastp reference database
    seqdb={}
    for record in input1:
        seq=str(record.seq)
        if record.id not in seqdb:
            seqdb[record.id]=seq

    input2= open(blast_out,"r") # blastp output
    input3= open(fasta2,"r") # novel peptide tab txt
    output= open('categorized_' + blast_out,"w")


    blastout={}
    hits_dic={}
    for line in input2:
        row=line.strip().split("\t")
        qid=row[-2]
        sid=row[1]
        sseq=seqdb[sid]
        ident=row[2]
        peplen=int(row[3])
        mismatch=row[4]
        alignlen=int(row[6])-int(row[5])+1
        sstart=int(row[7])
        send=int(row[8])
        gap=row[11]
        evalue=float(row[-5])
        alignseq=row[-1]
        category="NA"
        single_sub_pos="NA"
        if sstart>3:
            Nterm_seq=sseq[sstart-4:sstart+2] #check up 3 amino acid before N-term of this peptide
        else:
            Nterm_seq=sseq[:sstart]

        if len(sseq)-send<3:
            Cterm_seq=sseq[send-1:]
        else:
            Cterm_seq=sseq[send-3:send+3]

        if SAAV:
            if alignlen==peplen:
                if float(ident)==100:
                    category="match to known SAAV"
                
                elif int(gap)==0 and int(mismatch)==1:
                    category="potential novel SAAV (maps to known SAAV with 1 aa mismatch)"
                    for i in range(peplen):
                        if qid[i]!=alignseq[i]:
                            single_sub_pos=str(i+1)

                elif int(gap)==1 and int(mismatch)==0:
                    category="potential novel SAAV (maps to known SAAV with 1 aa insertion)"
                else:
                    category="potential novel SAAV (maps to known SAAV with more than 2 mismatched aa)"
            elif peplen-alignlen==1 and float(ident)==100:
                category="potential novel SAAV (maps to known SAAV with 1 aa deletion)"

            else:
                category="potential novel SAAV (maps to known SAAV with more than 2 mismatched aa)"
        
        else:
            if alignlen==peplen:
                if float(ident)==100:
                    category="match to known protein"
                
                elif int(gap)==0 and int(mismatch)==1:
                    category="map to known protein with 1 aa mismatch"
                    for i in range(peplen):
                        if qid[i]!=alignseq[i]:
                            single_sub_pos=str(i+1)

                elif int(gap)==1 and int(mismatch)==0:
                    category="map to known protein with 1 aa insertion"
                else:
                    category="potential novelpep (map to known protein with more than 2 mismatched aa)"
            elif peplen-alignlen==1 and float(ident)==100:
                category="map to known protein with 1 aa deletion"

            else:
                category="potential novelpep (map to known protein with more than 2 mismatched aa)"

        if qid not in hits_dic:
            hits_dic[qid]=evalue
            blastout[qid]=[category,sid,ident,peplen,single_sub_pos,Nterm_seq,alignseq,Cterm_seq,alignlen,mismatch,gap]
        else:
            if evalue<hits_dic[qid]:
                hits_dic[qid]=evalue
                blastout[qid]=[category,sid,ident,peplen,single_sub_pos,Nterm_seq,alignseq,Cterm_seq,alignlen,mismatch,gap]

    #header=input3.readline().strip().split("\t")

    header=["Query","blastp_category","blastp_match","identity","peplen","sub_pos","Nterm-seq(3aa)","aligned_seq","Cterm-seq(3aa)","alignlen","mismatch","gap"]
    output.write("\t".join(header)+"\n")

    for line in input3:
        row=line.strip().split("\t")
        peptide=row[0]
        if peptide in blastout:
            results=blastout[row[0]]
            newrow=row+results
            output.write("\t".join(map(str,newrow))+"\n")
        
        else:
            if SAAV:
                newrow=row+["potential novel SAAV (no match to known SAAV in dbSAP)","NA","NA","NA","NA","NA","NA","NA","NA","NA","NA"]
                output.write("\t".join(map(str,newrow))+"\n")
            else:
                newrow=row+["potential novelpep (no match to known protein found)","NA","NA","NA","NA","NA","NA","NA","NA","NA","NA"]
                output.write("\t".join(map(str,newrow))+"\n")


    input2.close()
    input3.close()
    output.close()
   
    #adding headers to the novel blast output
    header_names = ["qseqid", "sseqid", "pident", "qlen", "mismatch", "qstart", "qend", "sstart", "send", "evalue", "bitscore", "gaps", "qseq", "sseq"]
    novel_file = pd.read_table('categorized_' + blast_out, sep ='\t', names = header_names)
    header_names = ["qseqid", "sseqid", "pident", "qlen", "mismatch", "qstart", "qend", "sstart", "send", "evalue", "bitscore", "gaps", "qseq", "sseq"]
    blast_file = pd.read_table(blast_out, sep ='\t', names = header_names)
    
    #to check if any keys are missing in the dictionary seqdb[]
    # Change return statements to use logging
    missing = blast_file[~blast_file["sseqid"].isin(seqdb.keys())]
    if (len(missing) == 0):
        logger.info(f"\tall is good for {blast_out}")
        return True
    else:
        logger.error('dictionary error')
        return False

def get_pos(substring, string):
# Find the starting position
    start_position = string.find(substring)

    if start_position != -1:
        # Calculate the ending position
        end_position = start_position + len(substring) - 1
        return(start_position, end_position)


#fn that gets matches and position of matches from peptide_list to fasta
def find_matches(peptide, db, query):
    search_6ft = {}
    pos_6ft = {}
    for string in peptide:
        search_6ft[string] = []
        for key, value in db.items():
            if string in value:
                if (query == '6FT'):
                    combi = (key, get_pos(string,value))
                    search_6ft[string].append(combi)
                elif (query == 'PCPS'):
                    search_6ft[string].append(key)
                else:
                    logger.error("\tspecify find matches query: 6FT or PCPS")
                    return None
    unmatched = []
    #fn for filtering the 6FT matched dictionary
    def my_filtering_function(pair):
        key, value = pair
        if value == []:
            return False  # filter pair out of the dictionary
        else:
            return True  # keep pair in the filtered dictionary
        
    matched = dict(filter(my_filtering_function, search_6ft.items()))
    positions = dict(filter(my_filtering_function, pos_6ft.items()))

    for key, value in search_6ft.items():
        if (value == []):
            unmatched.append(key)
    if (len(matched) + len(unmatched) == len(search_6ft)):
        return (matched, unmatched)
    else:
        logger.error("\terror in separating matches")
        return None
    
def split_string(s):
    splits_dict = {}
    # Ensure the string is long enough to be split into two parts each with at least `min_length` characters
    if len(s) < 6:
        return splits_dict
    
    for i in range(3, len(s) - 2):
        part1 = s[:i]
        part2 = s[i:]
        if len(part2) >= 3:
            splits_dict[part1] = part2
    return splits_dict

def PCPS(input_file, db_canonical_fasta):
        
    input1= SeqIO.parse(db_canonical_fasta,"fasta") 
    seqdb={}
    for record in input1:
            seq=str(record.seq)
            if record.id not in seqdb:
                    seqdb[record.id]=seq
    
    input = open(input_file, "r")
    output = open('cis_PCPS', "w")
    output_2 = open('trans_PCPS', "w")
    for line in input:
            row=line.strip().split("\t")
            pep = row[0]
            split_combinations = split_string(pep)
            part1_list = split_combinations.keys()
            part2_list = split_combinations.values()
            match_1, unmatch_1 = find_matches(part1_list, seqdb, 'PCPS')
            match_2, unmatch_2 = find_matches(part2_list, seqdb, 'PCPS')
            row = []
            for key,value in match_1.items():
                    for pro_1 in match_1[key]:
                            if (split_combinations[key] in match_2.keys()):
                                    for pro_2 in match_2[split_combinations[key]]:
                                            if pro_1 == pro_2:
                                                result = (f"{key}|{split_combinations[key]}" + "\t" + f"{pro_2}")
                                                output.write(str(pep) + "\t" + result + "\n")
                                            else:
                                                result_2 = (f"{key}|{split_combinations[key]}" + "\t" + f"{pro_1}:{pro_2}")
                                                output_2.write(str(pep) + "\t" + result_2 + "\n")
    input.close()
    output.close()
    output_2.close()
    logger.info("\tseparate files cis_PCPS and trans_PCPS created for spliced peptides")
    return True


#
##
###
#### main funciton starts here


def setup_logging(output_file_path):
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    logger.addHandler(ch)

    log_path = os.path.join(output_file_path, "pipeline.log")
    fh = logging.FileHandler(log_path, mode="w")
    fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    logger.addHandler(fh)

    return logger

def parse_args():
    parser = argparse.ArgumentParser(description="Immunosearch for filtering and classifying MHC peptides")

    parser.add_argument("-w", "--work_dir", required=True, help="working direcrtory where runtime files are created")
    parser.add_argument("-i", "--input_file_path", required=True, help="Input xlsx file with PEAKS search result in sheet 1 and gibbs_clustering output in sheet 2")
    parser.add_argument("-m", "--MHC_class", required=True, help="MHC class")
    parser.add_argument("-g", "--gibbs_cluster", required=True, type=lambda s: [int(x) for x in s.split(',')],
                        help="comma-separated list of Gibbs cluster values to select")
    parser.add_argument("-o", "--output_file_path", required=True, help="Output directory")
    parser.add_argument("-f", "--file_name", required=True, help="xlsx file name")
    parser.add_argument("-d", "--db_path", required=True, help="path to databse folder with all databse files")
    parser.add_argument("-c", "--cleanup", action="store_true", help="flag to cleanup all files in the working directory, use to cleanup files from previous runs. !!DELETES all files in working directory, does not delete folders!!")

    return parser.parse_args()

def main():
    args = parse_args()

    # Expand paths
    work_dir = os.path.expanduser(args.work_dir)
    input_file_path = os.path.expanduser(args.input_file_path)
    output_file_path = os.path.expanduser(args.output_file_path)
    db_path = os.path.expanduser(args.db_path)
    MHC_class = args.MHC_class
    gibbs_cluster = args.gibbs_cluster
    cleanup = args.cleanup
    file_name = args.file_name

    try:
        os.makedirs(output_file_path, exist_ok=True)
        print(f"output directory created at {output_file_path}")
    except OSError as e:
        print(f"failed to create output directory {e}")
        sys.exit(1)

    logger = setup_logging(output_file_path)
    
    try:
        os.makedirs(work_dir, exist_ok=True)
        print(f"working directory created at {work_dir}")
    except OSError as e:
        print(f"failed to create working directory {e}")
        sys.exit(1)
    
    # Log and validate
    logger.info("Commencing Immunosearch...")
    logger.info(f"Working directory: {work_dir}")
    logger.info(f"Input file: {input_file_path}")
    logger.info(f"MHC Class: {MHC_class}")
    logger.info(f"Gibbs clusters selected: {gibbs_cluster}")
    logger.info(f"Cleanup enabled: {cleanup}")

    if cleanup:
        logger.info(f"Cleaning up files from previous runs in {work_dir}")
        for item in os.listdir(work_dir):
            item_path = os.path.join(work_dir, item)
            if os.path.isfile(item_path):
                try:
                    os.remove(item_path)
                    logger.info(f"\tremoved file {item}")
                except Exception as e:
                    logger.warning(f"\tfailed to remove {item} {e}")

    # Change to working directory
    os.chdir(work_dir)
    logger.info(f"Loading into working directory at {work_dir}")
    
    # peptide length 8-11, removing gibbs junk and DB matched
    all_data = pd.ExcelFile(input_file_path)
    data = pd.read_excel(all_data, file_name)
    gibbs = pd.read_excel(all_data, 'gibbs_clustering')
    logger.info("filtering peptides found by DB search")
    
    data = data[data["Found By"] != 'DB Search']
    if (len(gibbs_cluster) != 0):
        logger.info(f"Selected gibbs clusters based on input, gibbs clusters {gibbs_cluster}")
        gibbs = gibbs[gibbs["Gn"].isin(gibbs_cluster)]
    else:
        logger.info("selecting all gibbs clusters")
        gibbs = gibbs
    
    if (MHC_class == '1'):
        logger.info("selecting peptides with length between 8 and 11 AA")
        gibbs = gibbs[gibbs["Sequence"].str.len().between(8,11)]
    elif (MHC_class == '2'):
        logger.info("selecting peptides with length between 12 and 17 AA")
        gibbs = gibbs[gibbs["Sequence"].str.len().between(12,17)]
    elif (MHC_class == "E"):
        logger.info("selecting peptides with length between 8 and 15 AA")
        gibbs = gibbs[gibbs["Sequence"].str.len().between(8,15)]
    else:
        logger.warning("!!! not filtered by length")

    data = data[data["Peptide"].isin(gibbs["Sequence"])]
    list = pd.unique(data["Peptide"])
    pep = list.tolist()
    logger.info("generating lists and files for further analysis")
    create_files(pep,'peptides')


    #blast against known HLA peptides
    if(len(pep) != 0):

        if(len(pep) == 1): #coz I'm a grammar Nazi :)
            logger.info(f"{len(pep)} peptide being searched for known HLAs")
        else:
            logger.info(f"{len(pep)} peptides being searched for known HLAs")
        
        HLA_blastp_query_file = os.path.join(work_dir, "peptides.fasta")
        db_HLA_blast_db = os.path.join(db_path, "APD_Hs_all")
        HLA_blastp_output_file = os.path.join(work_dir, "HLA_blast_out")
                
        subprocess.run(
            ["blastp",
            "-task", "blastp-short",
            "-query", HLA_blastp_query_file,
            "-db", db_HLA_blast_db,
            "-out", HLA_blastp_output_file,
            "-evalue", "10.0",
            "-outfmt", "6 qseqid sseqid pident qlen mismatch qstart qend sstart send evalue bitscore gaps qseq sseq"],
            check=True)
        
        #subprocess.run("blastp -task blastp-short -query peptides.fasta -db db/APD_Hs_all -out HLA_blast_out -evalue 10.0 -outfmt \"6 qseqid sseqid pident qlen mismatch qstart qend sstart send evalue bitscore gaps qseq sseq\"", shell=True)

        #parsing and catergorizing HLA_blast output
        logger.info("\treading output")

        db_HLA_fasta = os.path.join(db_path, "APD_Hs_all.fasta")
        
        #parse_categorize('db/APD_Hs_all.fasta', 'HLA_blast_out', 'peptides_2.fasta')
        parse_categorize(db_HLA_fasta, 'HLA_blast_out', 'peptides_2.fasta')

        #known HLA
        logger.info("\tgenerating lists and files for further analysis")
        
        output_HLA = pd.read_table('categorized_HLA_blast_out')
        known = output_HLA[output_HLA["blastp_category"] == 'match to known protein']
        list = known["Query"]
        pep = list.to_list()

        if (len(pep) != 0):
            logger.info("\tknown HLA found")
            ofile = open("known_HLA.fasta", "w")

            for i in range(len(pep)):
                ofile.write('>' + '\n' + pep[i] + '\n')
            ofile.close()
        else:
            logger.info("\tno known HLA found")

        known = output_HLA[output_HLA["blastp_category"] != 'match to known protein']
        list = known["Query"] 
        pep = list.to_list()
    else:
        logger.info("No peptides to search for known HLAs")

    
    if (len(pep) != 0): #to prevent blast with empty query
        
        create_files(pep, 'to_blastp')

        if (len(pep) == 1):
            logger.info(f"{len(pep)} peptide being searched for human canonical proteins")
        else:
            logger.info(f"{len(pep)} peptides being searched for human canonical proteins")
        
        canonical_blastp_query_file = os.path.join(work_dir, "to_blastp.fasta")
        db_canonical_blast_db = os.path.join(db_path, "human_canonical")
        canonical_blastp_output_file = os.path.join(work_dir, "blastp_out_human_canonical")
        #blast all proteins against human canonical proteins
        subprocess.run(
            ["blastp",
            "-task", "blastp-short",
            "-query", canonical_blastp_query_file,
            "-db", db_canonical_blast_db,
            "-out", canonical_blastp_output_file,
            "-evalue", "10.0",
            "-outfmt", "6 qseqid sseqid pident qlen mismatch qstart qend sstart send evalue bitscore gaps qseq sseq"],
            check=True)
        #subprocess.run("blastp -task blastp-short -query to_blastp.fasta -db db/human_canonical -out blastp_out_human_canonical -evalue 10.0 -outfmt \"6 qseqid sseqid pident qlen mismatch qstart qend sstart send evalue bitscore gaps qseq sseq\"", shell=True)
        
        logger.info("\treading output")
        db_canonical_fasta = os.path.join(db_path, "human_canonical.fasta")
        #parse_categorize('db/human_canonical.fasta', 'blastp_out_human_canonical', 'to_blastp_2.fasta')
        parse_categorize(db_canonical_fasta, 'blastp_out_human_canonical', 'to_blastp_2.fasta')
        output  = pd.read_table('categorized_blastp_out_human_canonical')
        
        #preparing fasta files of proteins to search 6FT database
        
        to_6ft = output[(output["blastp_category"] != 'map to known protein with 1 aa mismatch') & (output["blastp_category"] != 'match to known protein')]
        list = to_6ft["Query"] 
        pep = list.to_list()

        create_files(pep, 'to_6ft')
       
        #preparing fasta files of proteins with 1 AA mismatch
        mismatched = output[output["blastp_category"] == 'map to known protein with 1 aa mismatch']
        list = mismatched["Query"] 
        pep = list.to_list()

    else:
        logger.warning("blastp input empty")

    logger.info("\tgenerating lists and files for further analysis")
    
    
    if (len(pep) != 0): #to prevent blast with empty query

        create_files(pep,'SAAV')
        
        if (len(pep) == 1):  #more grammar Nazi
            logger.info(f"{len(pep)} peptide being searched for Single Amino Acid Variants")
        else:
            logger.info(f"{len(pep)} peptides being searched for Single Amino Acid Variants")
            
        SAAV_blastp_query_file = os.path.join(work_dir, "SAAV.fasta")
        db_SAAV_blast_db = os.path.join(db_path, "sap_db")
        SAAV_blastp_output_file = os.path.join(work_dir, "blast_out_SAAV")
        #blastp against db_SAP (single amino acids polymorphisms)
        subprocess.run(
            ["blastp",
            "-task", "blastp-short",
            "-query", SAAV_blastp_query_file,
            "-db", db_SAAV_blast_db,
            "-out", SAAV_blastp_output_file,
            "-evalue", "10.0",
            "-outfmt", "6 qseqid sseqid pident qlen mismatch qstart qend sstart send evalue bitscore gaps qseq sseq"],
            check=True)        
        

        logger.info("\treading output")
        db_SAAV_fasta = os.path.join(db_path, "sap_db.fa")
        parse_categorize(db_SAAV_fasta , 'blast_out_SAAV', 'SAAV_2.fasta', SAAV = True)
        


        output = pd.read_table("categorized_blast_out_SAAV")
        not_SNP = output[output["blastp_category"] != 'match to known SAAV']
        pep = not_SNP["Query"]
        pep = pep.to_list()

        ofile = open("not_SAAV.fasta", "w")

        for i in range(len(pep)):
            ofile.write('>' + '\n' + pep[i] + '\n')
        ofile.close()

    else:
        logger.info("no potential SAAVs, proceeding to 6FT search")

    #searches peptides in 6FT db
    
    pep = to_6ft["Query"]

    if(len(pep) != 0):

        if(len(pep) == 1):
            
            logger.info(f"{len(pep)} peptide being searched in the six frame translated human genome")
            

            seqkit_query_file = os.path.join(work_dir, "to_6ft_2.fasta")
            db_human_6FT_fasta = os.path.join(db_path, "human_6FT_m.fasta")
            seqkit_output_file = os.path.join(work_dir, "6ft_out")

            cmd_seqkit = (f"seqkit grep --by-seq --ignore-case --threads 12 --seq-type protein "
                f"--pattern-file \"{seqkit_query_file}\" \"{db_human_6FT_fasta}\" > \"{seqkit_output_file}\"")
            logger.info(cmd_seqkit)
            subprocess.run(cmd_seqkit, shell=True, check=True)
            


            input1= SeqIO.parse('6ft_out',"fasta") # 6FT results to dict
            seqdb={}
            for record in input1:
                seq=str(record.seq)
                if record.description not in seqdb:
                    seqdb[record.description]=seq
            
            if (len(seqdb) == 0):
                logger.info("no match to 6FT")
                unmatched = pep
                matched = []
            else:
                matched = pep
                unmatched = []

        else:
            logger.info(f"{len(pep)} peptides being searched in the six frame translated human genome")

            seqkit_query_file = os.path.join(work_dir, "to_6ft_2.fasta")
            db_human_6FT_fasta = os.path.join(db_path, "human_6FT_m.fasta")
            seqkit_output_file = os.path.join(work_dir, "6ft_out")

            cmd_seqkit = (f"seqkit grep --by-seq --ignore-case --threads 12 --seq-type protein "
                f"--pattern-file \"{seqkit_query_file}\" \"{db_human_6FT_fasta}\" > \"{seqkit_output_file}\"")
            logger.info(cmd_seqkit)
            subprocess.run(cmd_seqkit, shell=True, check=True)


            logger.info("\twriting 6FT results to dictionary")

            input1= SeqIO.parse('6ft_out',"fasta") # 6FT results to dict
            seqdb={}
            for record in input1:
                seq=str(record.seq)
                if record.description not in seqdb:
                    seqdb[record.description]=seq
                    
            logger.info(f"\tnumber of matches from 6FT: {len(seqdb)}")

            input  = open('to_6ft_2.fasta', 'r') #list of peptides, has to be in a list
            pep = []
            for line in input:
                    pep = pep + line.strip().split("\t")

            logger.info("\tlocating and matching peptides to 6FT")

            matched, unmatched = find_matches(pep, seqdb, '6FT')

            if len(matched) == 0:
                logger.info("\tno peptides matched to 6ft")
            else:
                logger.info(f"\twriting {len(matched)} peptide matches to 6FT")
                with open('matches_to_6ft.csv','w') as f:
                    w = csv.writer(f)
                    w.writerow(["Peptide", "Sequence_loc"]) #header names
                    for key in matched:
                        sequence = matched[key]
                        w.writerow([key, sequence])

            logger.info(f"\twriting {len(unmatched)} unmatched peptides")
            create_files(unmatched, 'no_match_6ft')

    else:
        logger.info("no peptides to search in 6FT")
    

    #searching for proteosome catalyzed peptide spliced 
    
    logger.info(f"{len(unmatched)} peptides being searched for proteosome catalyzed peptide spliced variants")

    PCPS("no_match_6ft_2.fasta", db_canonical_fasta)
    
    
    #saving output files
    SAAV_out = pd.read_table('categorized_blast_out_SAAV', sep = "\t")
    Sixframe_notmatched = pd.read_table('no_match_6ft_2.fasta', sep = "\t", names = ["Peptides"])
    cis_PCPS = pd.read_table("cis_PCPS", names= ["Peptide", "Spliced_peptide", "Protein_origin"])
    output_HLA = pd.read_table('categorized_HLA_blast_out')
    known_HLA = output_HLA[output_HLA["blastp_category"] == 'match to known protein']
    #getting peptide loci

    # Load your data
    Sixframe_out = pd.read_csv('matches_to_6ft.csv')

    # Updated function: now includes frame info in the output
    def compute_locus_for_pair(entry_str, aa_pair):
        m_chr   = re.search(r'chromosome (\w+)', entry_str)
        m_beg   = re.search(r'begin=(\d+)', entry_str)
        m_end   = re.search(r'end=(\d+)',   entry_str)
        m_frame = re.search(r'frame=([+-]?\d+)', entry_str)

        chrom   = m_chr.group(1) if m_chr else None
        prot_b  = int(m_beg.group(1)) if m_beg else None
        prot_e  = int(m_end.group(1)) if m_end else None
        frame   = int(m_frame.group(1)) if m_frame else None

        s, e = aa_pair
        if frame > 0:
            g_start = prot_b + 3*(s - 1)
            g_end   = prot_b + 3*e   - 1
        else:
            g_start = prot_e - (3*e   - 1)
            g_end   = prot_e -  3*(s - 1)

        return f"chr{chrom}:{g_start}_{g_end}_frame={frame}"

    # Extract and join all loci with frame info
    def all_loci_from_cell(cell):
        pairs = ast.literal_eval(cell)
        loci = [
            compute_locus_for_pair(entry_str, aa_pair)
            for entry_str, aa_pair in pairs
        ]
        return ";".join(loci)

    # Apply to your data (replace 'ColumnName' with the correct column name)
    Sixframe_out['loci'] = Sixframe_out['Sequence_loc'].apply(all_loci_from_cell)
    
    
    #writing data to excel file
    if (len(mismatched) == 0 & len(matched) == 0 & len(unmatched) == 0):
        logger.info("No significant peptides found")
        
    else:
        os.chdir(output_file_path)
        logger.info("creating excel file with results")
        with pd.ExcelWriter(file_name + '_immuno_search_out.xlsx', engine='openpyxl') as writer:
            # Write each DataFrame to a different sheet
            if (len(known_HLA) != 0):
                known_HLA.to_excel(writer, sheet_name= "known_HLA", index= False)
            if (len(mismatched) != 0):
                SAAV_out.to_excel(writer, sheet_name='Single_AA_variants', index=False)
            if (len(matched) != 0):
                Sixframe_out.to_excel(writer, sheet_name='Matches_to_six_frame', index=False)
            if (len(unmatched) != 0):
                Sixframe_notmatched.to_excel(writer, sheet_name='Six_frame_non_matched', index=False)
            if (len(cis_PCPS) != 0):
                cis_PCPS.to_excel(writer, sheet_name= "cis_PCPS", index= False)

            #removes borders in excel headers
            for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    
                    # Get the dimensions of the sheet
                    max_row = worksheet.max_row
                    max_col = worksheet.max_column
                    
                    # Create a no-border style
                    no_border = Border(left=Side(style=None), 
                                    right=Side(style=None),
                                    top=Side(style=None),
                                    bottom=Side(style=None))
                    
                    # Apply no border to header row (row 1)
                    for col in range(1, max_col + 1):
                        cell = worksheet.cell(row=1, column=col)
                        cell.border = no_border
                    
                    # Auto-size columns based on content
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = get_column_letter(column[0].column)
                            
                            # Find the longest content in each column
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            
                            # Add some padding
                            adjusted_width = max_length + 2
                            
                            # Set a reasonable maximum width to prevent extremely wide columns
                            if adjusted_width > 50:
                                adjusted_width = 50
                            
                            # Set the column width
                            worksheet.column_dimensions[column_letter].width = adjusted_width

        os.chdir(work_dir)
        logger.info(f"returning back to work directory at {work_dir}")
        logger.info(f"search and classification done, file saved at {output_file_path}{file_name}_immuno_search_out.xlsx")

if __name__ == "__main__":
    main()